from openpyxl import Workbook
import os
import pandas as pd

def extract_solution_quality(input_file, output_file):

    #we create a matrix that holds the full information of the solution
    #the keys of the solution matrix are the instance_names
    instance_list = []
    window_size_list = []
    runs_per_window_list = []

    deadhead_matrix = {}
    uncovered_task_matrix = {}
    runtime_matrix = {}
    with open(input_file, "r", encoding="utf-8") as infile:
         with open(output_file, "w", encoding="utf-8") as outfile:
            for line in infile:
                if line.startswith("Instance:"):
                    parts = line.split()
                    instance_name = parts[1].replace(".tsv","")
                    if instance_name not in instance_list:
                        instance_list.append(instance_name)
                    window_size = parts[5]
                    if window_size not in window_size_list:
                        window_size_list.append(window_size)
                    runs_per_window = parts[10]
                    if runs_per_window not in runs_per_window_list:
                        runs_per_window_list.append(runs_per_window)
                    #outfile.write(line)
                if line.startswith("After performing"):
                    parts = line.split()
                    deadheadings = str(int(int(parts[9].replace(",",""))/999))
                    uncovered_tasks = parts[13]
                    deadhead_matrix[(instance_name, window_size, runs_per_window)] = deadheadings
                    uncovered_task_matrix[(instance_name, window_size, runs_per_window)] = uncovered_tasks
                    #outfile.write(f"#deadheadings: {deadheadings} | uncovered: {uncovered_tasks} \n")
                if line.startswith("The total"):
                    parts = line.split()
                    runtime = str(round(float(parts[4]),1))
                    runtime_matrix[(instance_name, window_size, runs_per_window)] = runtime
                    #outfile.write(f"runtime: {runtime} \n")

            print(runtime_matrix)
            for window_size in window_size_list:
                outfile.write(f"window_size {window_size} \n")
                for runs_per_window in runs_per_window_list:
                    if runs_per_window == "1":
                        outfile.write(f"& {runs_per_window} ")
                    else:
                        outfile.write(f"& & {runs_per_window} ")
                    for instance_name in instance_list:
                        if (instance_name, window_size, runs_per_window) in uncovered_task_matrix.keys() and (instance_name, window_size, runs_per_window) in deadhead_matrix.keys() and (instance_name, window_size, runs_per_window) in runtime_matrix.keys():
                            outfile.write(f"& {uncovered_task_matrix[(instance_name, window_size, runs_per_window)]} & {deadhead_matrix[(instance_name, window_size, runs_per_window)]} & {runtime_matrix[(instance_name, window_size, runs_per_window)]} ")
                        else:
                            outfile.write(
                                f"& - & - & - ")
                    outfile.write(" \\\ \n")

    infile.close()
    outfile.close()

def extract_solution_quality_sideways_table(input_file, output_file):
    instance_list = []
    window_size_list = []
    runs_per_window_list = []
    method_list = []
    min_loco_knowledge_list = []
    min_section_knowledge_list = []
    max_deadhead_duration_list = []
    random_iteration_list = []

    deadhead_matrix = {}
    uncovered_task_matrix = {}
    runtime_matrix = {}

    # Parse the input file
    with open(input_file, "r", encoding="utf-8") as infile:
        for line in infile:
            if line.startswith("Instance:"):
                parts = line.split()
                instance_name = parts[1].replace(".tsv", "")
                if instance_name not in instance_list:
                    instance_list.append(instance_name)

                window_size = parts[5]
                if window_size not in window_size_list:
                    window_size_list.append(window_size)

                runs_per_window = parts[10]
                if runs_per_window not in runs_per_window_list:
                    runs_per_window_list.append(runs_per_window)

                method = parts[13]
                if method not in method_list:
                    method_list.append(method)

                min_loco_knowledge = parts[18]
                if min_loco_knowledge not in min_loco_knowledge_list:
                    min_loco_knowledge_list.append(min_loco_knowledge)

                min_section_knowledge = parts[23]
                if min_section_knowledge not in min_section_knowledge_list:
                    min_section_knowledge_list.append(min_section_knowledge)

                max_deadhead_duration = parts[28]
                if max_deadhead_duration not in max_deadhead_duration_list:
                    max_deadhead_duration_list.append(max_deadhead_duration)

                random_iteration = parts[32]
                if random_iteration not in random_iteration_list:
                    random_iteration_list.append(random_iteration)

            elif line.startswith("After performing"):
                parts = line.split()
                deadheadings = str(int(int(parts[9].replace(",", ""))))
                uncovered_tasks = parts[13]

                deadhead_matrix[(instance_name, window_size, runs_per_window, method, min_loco_knowledge, min_section_knowledge, max_deadhead_duration, random_iteration)] = deadheadings
                uncovered_task_matrix[(instance_name, window_size, runs_per_window, method, min_loco_knowledge, min_section_knowledge, max_deadhead_duration, random_iteration)] = uncovered_tasks

            elif line.startswith("The total"):
                parts = line.split()
                runtime = str(round(float(parts[4]), 1))
                runtime_matrix[(instance_name, window_size, runs_per_window, method, min_loco_knowledge, min_section_knowledge, max_deadhead_duration, random_iteration)] = runtime

    # Sort lists for consistent order
    instance_list.sort()
    window_size_list.sort()
    runs_per_window_list.sort(key=lambda x: int(x))

    ###########################################################
    avg_deadhead_matrix = {}
    avg_uncovered_task_matrix = {}
    avg_runtime_matrix = {}

    #calculate average over the random iterations
    for instance_name in instance_list:
        for window_size in window_size_list:
            for runs_per_window in runs_per_window_list:
                for method in method_list:
                    for min_loco_knowledge in min_loco_knowledge_list:
                        for min_section_knowledge in min_section_knowledge_list:
                            for max_deadhead_duration in max_deadhead_duration_list:
                                deadheads = 0
                                uncovered_tasks = 0
                                runtime = 0

                                at_least_one_iteration_found = True
                                for random_iteration in random_iteration_list:
                                    key = (instance_name, window_size, runs_per_window, method, min_loco_knowledge, min_section_knowledge, max_deadhead_duration, random_iteration)
                                    if key in deadhead_matrix:
                                        deadheads += float(deadhead_matrix[key])
                                    else:
                                        at_least_one_iteration_found = False
                                    if key in uncovered_task_matrix:
                                        uncovered_tasks += float(uncovered_task_matrix[key])
                                    else:
                                        at_least_one_iteration_found = False
                                    if key in runtime_matrix:
                                        runtime += float(runtime_matrix[key])
                                    else:
                                        at_least_one_iteration_found = False
                                if at_least_one_iteration_found == True:
                                    avg_deadhead_matrix[(instance_name, window_size, runs_per_window, method, min_loco_knowledge, min_section_knowledge, max_deadhead_duration)] = str(deadheads / len(random_iteration_list))
                                    avg_uncovered_task_matrix[(instance_name, window_size, runs_per_window, method, min_loco_knowledge, min_section_knowledge, max_deadhead_duration)] = str(uncovered_tasks / len(random_iteration_list))
                                    avg_runtime_matrix[(instance_name, window_size, runs_per_window, method, min_loco_knowledge, min_section_knowledge, max_deadhead_duration)] = str(round(runtime / len(random_iteration_list),2))

    ###########################################################

    #We write several tables, one for each combination of section and locomotive knowledge and also one for each max_deadhead_duration
    for method in method_list:
        for min_loco_knowledge in min_loco_knowledge_list:
            for min_section_knowledge in min_section_knowledge_list:
                for max_deadhead_duration in max_deadhead_duration_list:
                    # Write LaTeX table rows
                    with open(output_file.replace(".log", "_"+str(method)+"_"+str(min_loco_knowledge)+"_"+str(min_section_knowledge)+"_"+str(max_deadhead_duration)+".log"), "w", encoding="utf-8") as outfile:
                        for inst in instance_list:
                            outfile.write(f"{inst} & ")
                            # Insert characteristics placeholder, can be modified to actual values
                            outfile.write(r"\begin{tabular}{c}" + "\n")
                            outfile.write(r"\#tasks = ...\\")
                            outfile.write(r"\#dh = ...\\")
                            outfile.write(r"dur = ... min\\")
                            outfile.write(r"horizon = ... min")
                            outfile.write(r"\end{tabular} ")

                            # Go through DP windows and iterations
                            for window in ["150", "300", "450"]:
                                for it in ["1", "5", "10"]:
                                    key = (inst, window, it, "DP", min_loco_knowledge, min_section_knowledge, max_deadhead_duration)
                                    if key in avg_uncovered_task_matrix:
                                        outfile.write(
                                            f"& {avg_uncovered_task_matrix[key]} & {avg_deadhead_matrix[key]} & {avg_runtime_matrix[key]} "
                                        )
                                    else:
                                        outfile.write("& - & - & - ")

                            # DP full horizon
                            for window in ["10800"]:
                                for it in ["1", "5", "10"]:
                                    key = (inst, window, it, "DP", min_loco_knowledge, min_section_knowledge, max_deadhead_duration)
                                    if key in avg_uncovered_task_matrix:
                                        outfile.write(f"& {avg_uncovered_task_matrix[key]} & {avg_deadhead_matrix[key]} & {avg_runtime_matrix[key]} ")
                                    else:
                                        outfile.write("& - & - & - ")

                            # Model columns (150, 300, full)
                            for window in ["150", "300"]:
                                key = (inst, window, "1", "model", min_loco_knowledge, min_section_knowledge, max_deadhead_duration)  # adjust keys if needed
                                if key in avg_uncovered_task_matrix:
                                    outfile.write(f"& {avg_uncovered_task_matrix[key]} & {avg_deadhead_matrix[key]} & {avg_runtime_matrix[key]} ")
                                else:
                                    outfile.write("& - & - & - ")

                            outfile.write(r"\\")
                            outfile.write("\n")

def extract_solution_quality_sideways_table_excel(input_file, output_file):
    instance_list = []
    window_size_list = []
    runs_per_window_list = []
    method_list = []
    min_loco_knowledge_list = []
    min_section_knowledge_list = []
    max_deadhead_duration_list = []
    random_iteration_list = []

    nr_deadhead_matrix = {}
    deadhead_matrix = {}
    uncovered_task_matrix = {}
    runtime_matrix = {}
    spared_driver_matrix = {}

    # Parse input
    if input_file.endswith(".log"):
        with open(input_file, "r", encoding="utf-8") as infile:
            for line in infile:
                if line.startswith("Instance:"):
                    parts = line.split()
                    instance_name = parts[1].replace(".tsv", "")
                    if instance_name not in instance_list:
                        instance_list.append(instance_name)

                    window_size = parts[5]
                    if window_size not in window_size_list:
                        window_size_list.append(window_size)

                    runs_per_window = parts[10]
                    if runs_per_window not in runs_per_window_list:
                        runs_per_window_list.append(runs_per_window)

                    method = parts[13]
                    if method not in method_list:
                        method_list.append(method)

                    min_loco_knowledge = parts[18]
                    if min_loco_knowledge not in min_loco_knowledge_list:
                        min_loco_knowledge_list.append(min_loco_knowledge)

                    min_section_knowledge = parts[23]
                    if min_section_knowledge not in min_section_knowledge_list:
                        min_section_knowledge_list.append(min_section_knowledge)

                    max_deadhead_duration = parts[28]
                    if max_deadhead_duration not in max_deadhead_duration_list:
                        max_deadhead_duration_list.append(max_deadhead_duration)

                    random_iteration = parts[32]
                    if random_iteration not in random_iteration_list:
                        random_iteration_list.append(random_iteration)

                elif line.startswith("After performing"):
                    parts = line.split()
                    deadheadings = str(int(int(parts[9].replace(",", ""))))
                    uncovered_tasks = parts[13]
                    nr_deadheads = parts[32]
                    spared_driver = parts[27]

                    nr_deadhead_matrix[(instance_name, window_size, runs_per_window, method,
                                     min_loco_knowledge, min_section_knowledge,
                                     max_deadhead_duration, random_iteration)] = nr_deadheads

                    deadhead_matrix[(instance_name, window_size, runs_per_window, method,
                                     min_loco_knowledge, min_section_knowledge,
                                     max_deadhead_duration, random_iteration)] = deadheadings

                    uncovered_task_matrix[(instance_name, window_size, runs_per_window, method,
                                           min_loco_knowledge, min_section_knowledge,
                                           max_deadhead_duration, random_iteration)] = uncovered_tasks

                    spared_driver_matrix[(instance_name, window_size, runs_per_window, method,
                                           min_loco_knowledge, min_section_knowledge,
                                           max_deadhead_duration, random_iteration)] = spared_driver

                elif line.startswith("The total"):
                    parts = line.split()
                    runtime = str(round(float(parts[4]), 1))
                    runtime_matrix[(instance_name, window_size, runs_per_window, method,
                                    min_loco_knowledge, min_section_knowledge,
                                    max_deadhead_duration, random_iteration)] = runtime

    # Sort for consistent order
    instance_list.sort()
    window_size_list.sort()
    runs_per_window_list.sort(key=lambda x: int(x))

    avg_nr_deadhead_matrix = {}
    avg_deadhead_matrix = {}
    avg_uncovered_task_matrix = {}
    avg_runtime_matrix = {}
    avg_spared_driver_matrix = {}

    # Compute averages
    for instance_name in instance_list:
        for window_size in window_size_list:
            for runs_per_window in runs_per_window_list:
                for method in method_list:
                    for min_loco_knowledge in min_loco_knowledge_list:
                        for min_section_knowledge in min_section_knowledge_list:
                            for max_deadhead_duration in max_deadhead_duration_list:
                                nr_deadheads = 0
                                deadheads = 0
                                uncovered_tasks = 0
                                runtime = 0
                                spared_driver = 0
                                valid = True

                                for random_iteration in random_iteration_list:
                                    key = (instance_name, window_size, runs_per_window, method,
                                           min_loco_knowledge, min_section_knowledge,
                                           max_deadhead_duration, random_iteration)

                                    if key not in deadhead_matrix or \
                                       key not in uncovered_task_matrix or \
                                       key not in runtime_matrix:
                                        valid = False
                                        break

                                    nr_deadheads += float(nr_deadhead_matrix[key])
                                    deadheads += float(deadhead_matrix[key])
                                    uncovered_tasks += float(uncovered_task_matrix[key])
                                    runtime += float(runtime_matrix[key])
                                    spared_driver += float(spared_driver_matrix[key])

                                if valid:
                                    main_key = (instance_name, window_size, runs_per_window,
                                                method, min_loco_knowledge,
                                                min_section_knowledge, max_deadhead_duration)

                                    avg_nr_deadhead_matrix[main_key] = nr_deadheads / len(random_iteration_list)
                                    avg_deadhead_matrix[main_key] = deadheads / len(random_iteration_list)
                                    avg_uncovered_task_matrix[main_key] = uncovered_tasks / len(random_iteration_list)
                                    avg_runtime_matrix[main_key] = round(runtime / len(random_iteration_list), 2)
                                    avg_spared_driver_matrix[main_key] = spared_driver / len(random_iteration_list)

    # Create Excel files instead of logs
    for method in method_list:
        for min_loco_knowledge in min_loco_knowledge_list:
            for min_section_knowledge in min_section_knowledge_list:
                for max_deadhead_duration in max_deadhead_duration_list:

                    filename = output_file.replace(
                        ".log",
                        f"_{method}_{min_loco_knowledge}_{min_section_knowledge}_{max_deadhead_duration}.xlsx"
                    )

                    wb = Workbook()
                    ws = wb.active
                    ws.title = "Results"

                    # Header row
                    headers = [
                        "Instance",
                        "Characteristics",
                    ]

                    # DP windows & iterations
                    for window in ["150", "300", "450"]:
                        for it in ["1", "5", "10"]:
                            headers += [
                                f"DP {window}/{it} - Uncov",
                                f"DP {window}/{it} - Nr. Deadheads",
                                f"DP {window}/{it} - Deadhead Meters",
                                f"DP {window}/{it} - Runtime",
                                f"DP {window}/{it} - Nr. Spared Drivers"
                            ]

                    # DP full
                    for window in ["10800"]:
                        for it in ["1", "5", "10"]:
                            headers += [
                                f"DP FH/{it} - Uncov",
                                f"DP FH/{it} - Nr. Deadheads",
                                f"DP FH/{it} - Deadhead Meters",
                                f"DP FH/{it} - Runtime",
                                f"DP FH/{it} - Nr. Spared Drivers"
                            ]

                    # Model runs
                    for window in ["150", "300"]:
                        headers += [
                            f"Model {window} - Uncov",
                            f"Model {window} - Nr. Deadheads",
                            f"Model {window} - Deadhead Meters",
                            f"Model {window} - Runtime",
                            f"Model {window} - Nr. Spared Drivers"
                        ]

                    # model full
                    for window in ["10800"]:
                        headers += [
                            f"Model FH - Uncov",
                            f"Model FH - Nr. Deadheads",
                            f"Model FH - Deadhead Meters",
                            f"Model FH - Runtime",
                            f"Model FH - Nr. Spared Drivers"
                        ]

                    ws.append(headers)

                    # Fill table rows
                    for inst in instance_list:
                        row = [
                            inst,
                            "#tasks=...; #dh=...; dur=...; horizon=..."
                        ]

                        for window in ["150", "300", "450"]:
                            for it in ["1", "5", "10"]:
                                key = (inst, window, it, "DP",
                                       min_loco_knowledge, min_section_knowledge,
                                       max_deadhead_duration)

                                if key in avg_uncovered_task_matrix:
                                    row += [
                                        avg_uncovered_task_matrix[key],
                                        avg_nr_deadhead_matrix[key],
                                        avg_deadhead_matrix[key],
                                        avg_runtime_matrix[key],
                                        avg_spared_driver_matrix[key]
                                    ]
                                else:
                                    row += ["-", "-", "-", "-", "-"]

                        # Full horizon DP
                        for window in ["10800"]:
                            for it in ["1", "5", "10"]:
                                key = (inst, window, it, "DP",
                                       min_loco_knowledge, min_section_knowledge,
                                       max_deadhead_duration)

                                if key in avg_uncovered_task_matrix:
                                    row += [
                                        avg_uncovered_task_matrix[key],
                                        avg_nr_deadhead_matrix[key],
                                        avg_deadhead_matrix[key],
                                        avg_runtime_matrix[key],
                                        avg_spared_driver_matrix[key]
                                    ]
                                else:
                                    row += ["-", "-", "-", "-", "-"]

                        # Model results
                        for window in ["150", "300"]:
                            key = (inst, window, "1", "model",
                                   min_loco_knowledge, min_section_knowledge,
                                   max_deadhead_duration)

                            if key in avg_uncovered_task_matrix:
                                row += [
                                    avg_uncovered_task_matrix[key],
                                    avg_nr_deadhead_matrix[key],
                                    avg_deadhead_matrix[key],
                                    avg_runtime_matrix[key],
                                    avg_spared_driver_matrix[key]
                                ]
                            else:
                                row += ["-", "-", "-", "-", "-"]

                        # Model results
                        for window in ["10800"]:
                            key = (inst, window, "1", "model",
                                   min_loco_knowledge, min_section_knowledge,
                                   max_deadhead_duration)

                            if key in avg_uncovered_task_matrix:
                                row += [
                                    avg_uncovered_task_matrix[key],
                                    avg_nr_deadhead_matrix[key],
                                    avg_deadhead_matrix[key],
                                    avg_runtime_matrix[key],
                                    avg_spared_driver_matrix[key]
                                ]
                            else:
                                row += ["-", "-", "-", "-", "-"]

                        ws.append(row)

                    wb.save(filename)

def extract_solution_quality_sideways_table_excel_incl_sparedrivers(input_file, output_file):
    instance_list = []
    window_size_list = []
    runs_per_window_list = []
    method_list = []
    min_loco_knowledge_list = []
    min_section_knowledge_list = []
    max_deadhead_duration_list = []
    fraction_spare_drivers_list = []
    random_iteration_list = []

    nr_deadhead_matrix = {}
    deadhead_matrix = {}
    uncovered_task_matrix = {}
    runtime_matrix = {}
    spared_driver_matrix = {}

    # Parse input
    if input_file.endswith(".log"):
        with open(input_file, "r", encoding="utf-8") as infile:
            for line in infile:
                if line.startswith("Instance:"):
                    parts = line.split()
                    instance_name = parts[1].replace(".tsv", "")
                    if instance_name not in instance_list:
                        instance_list.append(instance_name)

                    window_size = parts[5]
                    if window_size not in window_size_list:
                        window_size_list.append(window_size)

                    runs_per_window = parts[10]
                    if runs_per_window not in runs_per_window_list:
                        runs_per_window_list.append(runs_per_window)

                    method = parts[13]
                    if method not in method_list:
                        method_list.append(method)

                    min_loco_knowledge = parts[18]
                    if min_loco_knowledge not in min_loco_knowledge_list:
                        min_loco_knowledge_list.append(min_loco_knowledge)

                    min_section_knowledge = parts[23]
                    if min_section_knowledge not in min_section_knowledge_list:
                        min_section_knowledge_list.append(min_section_knowledge)

                    max_deadhead_duration = parts[28]
                    if max_deadhead_duration not in max_deadhead_duration_list:
                        max_deadhead_duration_list.append(max_deadhead_duration)

                    random_iteration = parts[32]
                    if random_iteration not in random_iteration_list:
                        random_iteration_list.append(random_iteration)

                    fraction_spare_drivers = parts[37]
                    if fraction_spare_drivers not in fraction_spare_drivers_list:
                        fraction_spare_drivers_list.append(fraction_spare_drivers)

                elif line.startswith("After performing"):
                    parts = line.split()
                    deadheadings = str(int(int(parts[9].replace(",", ""))))
                    uncovered_tasks = parts[13]
                    nr_deadheads = parts[32]
                    spared_driver = parts[27]

                    nr_deadhead_matrix[(instance_name, window_size, runs_per_window, method,
                                     min_loco_knowledge, min_section_knowledge,
                                     max_deadhead_duration, random_iteration, fraction_spare_drivers)] = nr_deadheads

                    deadhead_matrix[(instance_name, window_size, runs_per_window, method,
                                     min_loco_knowledge, min_section_knowledge,
                                     max_deadhead_duration, random_iteration, fraction_spare_drivers)] = deadheadings

                    uncovered_task_matrix[(instance_name, window_size, runs_per_window, method,
                                           min_loco_knowledge, min_section_knowledge,
                                           max_deadhead_duration, random_iteration, fraction_spare_drivers)] = uncovered_tasks

                    spared_driver_matrix[(instance_name, window_size, runs_per_window, method,
                                           min_loco_knowledge, min_section_knowledge,
                                           max_deadhead_duration, random_iteration, fraction_spare_drivers)] = spared_driver

                elif line.startswith("The total"):
                    parts = line.split()
                    runtime = str(round(float(parts[4]), 1))
                    runtime_matrix[(instance_name, window_size, runs_per_window, method,
                                    min_loco_knowledge, min_section_knowledge,
                                    max_deadhead_duration, random_iteration, fraction_spare_drivers)] = runtime

    # Sort for consistent order
    instance_list.sort()
    window_size_list.sort()
    runs_per_window_list.sort(key=lambda x: int(x))

    avg_nr_deadhead_matrix = {}
    avg_deadhead_matrix = {}
    avg_uncovered_task_matrix = {}
    avg_runtime_matrix = {}
    avg_spared_driver_matrix = {}

    # Compute averages
    for instance_name in instance_list:
        for window_size in window_size_list:
            for runs_per_window in runs_per_window_list:
                for method in method_list:
                    for min_loco_knowledge in min_loco_knowledge_list:
                        for min_section_knowledge in min_section_knowledge_list:
                            for max_deadhead_duration in max_deadhead_duration_list:
                                for fraction_spare_drivers in fraction_spare_drivers_list:
                                    nr_deadheads = 0
                                    deadheads = 0
                                    uncovered_tasks = 0
                                    runtime = 0
                                    spared_driver = 0
                                    valid = True

                                    for random_iteration in random_iteration_list:
                                        key = (instance_name, window_size, runs_per_window, method,
                                               min_loco_knowledge, min_section_knowledge,
                                               max_deadhead_duration, random_iteration, fraction_spare_drivers)

                                        if key not in deadhead_matrix or \
                                           key not in uncovered_task_matrix or \
                                           key not in runtime_matrix:
                                            valid = False
                                            break

                                        nr_deadheads += float(nr_deadhead_matrix[key])
                                        deadheads += float(deadhead_matrix[key])
                                        uncovered_tasks += float(uncovered_task_matrix[key])
                                        runtime += float(runtime_matrix[key])
                                        spared_driver += float(spared_driver_matrix[key])

                                    if valid:
                                        main_key = (instance_name, window_size, runs_per_window,
                                                    method, min_loco_knowledge,
                                                    min_section_knowledge, max_deadhead_duration, fraction_spare_drivers)

                                        avg_nr_deadhead_matrix[main_key] = nr_deadheads / len(random_iteration_list)
                                        avg_deadhead_matrix[main_key] = deadheads / len(random_iteration_list)
                                        avg_uncovered_task_matrix[main_key] = uncovered_tasks / len(random_iteration_list)
                                        avg_runtime_matrix[main_key] = round(runtime / len(random_iteration_list), 2)
                                        avg_spared_driver_matrix[main_key] = spared_driver / len(random_iteration_list)

    # Create Excel files instead of logs
    for method in method_list:
        for min_loco_knowledge in min_loco_knowledge_list:
            for min_section_knowledge in min_section_knowledge_list:
                for max_deadhead_duration in max_deadhead_duration_list:
                    for fraction_spare_drivers in fraction_spare_drivers_list:
                        filename = output_file.replace(
                            ".log",
                            f"_{method}_{min_loco_knowledge}_{min_section_knowledge}_{max_deadhead_duration}_{fraction_spare_drivers}.xlsx"
                        )

                        wb = Workbook()
                        ws = wb.active
                        ws.title = "Results"

                        # Header row
                        headers = [
                            "Instance",
                            "Characteristics",
                        ]

                        # DP windows & iterations
                        for window in ["150", "300", "450"]:
                            for it in ["1", "5", "10"]:
                                headers += [
                                    f"DP {window}/{it} - Uncov",
                                    f"DP {window}/{it} - Nr. Deadheads",
                                    f"DP {window}/{it} - Deadhead Meters",
                                    f"DP {window}/{it} - Runtime",
                                    f"DP {window}/{it} - Nr. Spared Drivers"
                                ]

                        # DP full
                        for window in ["10800"]:
                            for it in ["1", "5", "10"]:
                                headers += [
                                    f"DP FH/{it} - Uncov",
                                    f"DP FH/{it} - Nr. Deadheads",
                                    f"DP FH/{it} - Deadhead Meters",
                                    f"DP FH/{it} - Runtime",
                                    f"DP FH/{it} - Nr. Spared Drivers"
                                ]

                        # Model runs
                        for window in ["150", "300"]:
                            headers += [
                                f"Model {window} - Uncov",
                                f"Model {window} - Nr. Deadheads",
                                f"Model {window} - Deadhead Meters",
                                f"Model {window} - Runtime",
                                f"Model {window} - Nr. Spared Drivers"
                            ]

                        # model full
                        for window in ["10800"]:
                            headers += [
                                f"Model FH - Uncov",
                                f"Model FH - Nr. Deadheads",
                                f"Model FH - Deadhead Meters",
                                f"Model FH - Runtime",
                                f"Model FH - Nr. Spared Drivers"
                            ]

                        ws.append(headers)

                        # Fill table rows
                        for inst in instance_list:
                            row = [
                                inst,
                                "#tasks=...; #dh=...; dur=...; horizon=..."
                            ]

                            for window in ["150", "300", "450"]:
                                for it in ["1", "5", "10"]:
                                    key = (inst, window, it, "DP",
                                           min_loco_knowledge, min_section_knowledge,
                                           max_deadhead_duration, fraction_spare_drivers)

                                    if key in avg_uncovered_task_matrix:
                                        row += [
                                            avg_uncovered_task_matrix[key],
                                            avg_nr_deadhead_matrix[key],
                                            avg_deadhead_matrix[key],
                                            avg_runtime_matrix[key],
                                            avg_spared_driver_matrix[key]
                                        ]
                                    else:
                                        row += ["-", "-", "-", "-", "-"]

                            # Full horizon DP
                            for window in ["10800"]:
                                for it in ["1", "5", "10"]:
                                    key = (inst, window, it, "DP",
                                           min_loco_knowledge, min_section_knowledge,
                                           max_deadhead_duration, fraction_spare_drivers)

                                    if key in avg_uncovered_task_matrix:
                                        row += [
                                            avg_uncovered_task_matrix[key],
                                            avg_nr_deadhead_matrix[key],
                                            avg_deadhead_matrix[key],
                                            avg_runtime_matrix[key],
                                            avg_spared_driver_matrix[key]
                                        ]
                                    else:
                                        row += ["-", "-", "-", "-", "-"]

                            # Model results
                            for window in ["150", "300"]:
                                key = (inst, window, "1", "model",
                                       min_loco_knowledge, min_section_knowledge,
                                       max_deadhead_duration, fraction_spare_drivers)

                                if key in avg_uncovered_task_matrix:
                                    row += [
                                        avg_uncovered_task_matrix[key],
                                        avg_nr_deadhead_matrix[key],
                                        avg_deadhead_matrix[key],
                                        avg_runtime_matrix[key],
                                        avg_spared_driver_matrix[key]
                                    ]
                                else:
                                    row += ["-", "-", "-", "-", "-"]

                            # Model results
                            for window in ["10800"]:
                                key = (inst, window, "1", "model",
                                       min_loco_knowledge, min_section_knowledge,
                                       max_deadhead_duration, fraction_spare_drivers)

                                if key in avg_uncovered_task_matrix:
                                    row += [
                                        avg_uncovered_task_matrix[key],
                                        avg_nr_deadhead_matrix[key],
                                        avg_deadhead_matrix[key],
                                        avg_runtime_matrix[key],
                                        avg_spared_driver_matrix[key]
                                    ]
                                else:
                                    row += ["-", "-", "-", "-", "-"]

                            ws.append(row)

                        wb.save(filename)


def combine_excel_tables(input_folder, output_folder, suffix, sort_column="Instance"):
    """
    Combines all Excel files in input_folder that end with the given suffix,
    collapses multiple rows per 'Instance' keeping real values instead of '-',
    sorts the combined table by sort_column, and saves the result
    into output_folder.

    :param input_folder: folder containing the Excel files
    :param output_folder: folder to save the combined Excel
    :param suffix: suffix of files to combine, e.g. "_0.5_0.5_720.xlsx"
    :param sort_column: column name to sort by (default "Instance")
    """
    os.makedirs(output_folder, exist_ok=True)

    # Find matching files
    matching_files = [
        os.path.join(input_folder, f)
        for f in os.listdir(input_folder)
        if f.endswith(suffix) and os.path.isfile(os.path.join(input_folder, f))
    ]

    if not matching_files:
        print("No files found for suffix:", suffix)
        return

    print("Found files:")
    for f in matching_files:
        print("  -", f)

    # Read and append all tables
    combined_df = pd.concat(
        (pd.read_excel(f) for f in matching_files),
        ignore_index=True
    )

    # Collapse multiple rows per 'Instance'
    def combine_rows(group):
        # For each column, pick the first value that is not '-'
        return group.apply(lambda col: next((v for v in col if v != "-"), "-"))

    combined_df = combined_df.groupby("Instance", as_index=False).apply(combine_rows)

    # Sort by column
    if sort_column in combined_df.columns:
        combined_df = combined_df.sort_values(by=sort_column, ascending=True)
    else:
        print(f"Warning: Column '{sort_column}' not found. Skipping sort.")

    # Output filename in output folder
    output_file = os.path.join(
        output_folder,
        f"Combined{suffix}"
    )

    combined_df.to_excel(output_file, index=False)
    print(f"\nSaved combined, collapsed, and sorted file as: {output_file}")

def extract_solution_quality_section_and_loco_knowledge(input_file, output_file):

    #we create a matrix that holds the full information of the solution
    #the keys of the solution matrix are the instance_names
    instance_list = []
    window_size_list = []
    runs_per_window_list = []
    min_section_knowledge_list = []
    min_loco_knowledge_list = []

    deadhead_matrix = {}
    uncovered_task_matrix = {}
    runtime_matrix = {}
    with open(input_file, "r", encoding="utf-8") as infile:
         with open(output_file, "w", encoding="utf-8") as outfile:
            for line in infile:
                if line.startswith("Instance:"):
                    parts = line.split()
                    instance_name = parts[1].replace(".tsv","")
                    if instance_name not in instance_list:
                        instance_list.append(instance_name)
                    window_size = parts[5]
                    if window_size not in window_size_list:
                        window_size_list.append(window_size)
                    runs_per_window = parts[10]
                    if runs_per_window not in runs_per_window_list:
                        runs_per_window_list.append(runs_per_window)
                    min_loco_knowledge = parts[15]
                    if min_loco_knowledge not in min_loco_knowledge_list:
                        min_loco_knowledge_list.append(min_loco_knowledge)
                    min_section_knowledge = parts[20]
                    if min_section_knowledge not in min_section_knowledge_list:
                        min_section_knowledge_list.append(min_section_knowledge)
                    #outfile.write(line)
                if line.startswith("After performing"):
                    parts = line.split()
                    deadheadings = str(int(int(parts[9].replace(",",""))/999))
                    uncovered_tasks = parts[13]
                    deadhead_matrix[(instance_name, window_size, runs_per_window, min_loco_knowledge, min_section_knowledge)] = deadheadings
                    uncovered_task_matrix[(instance_name, window_size, runs_per_window, min_loco_knowledge, min_section_knowledge)] = uncovered_tasks
                    #outfile.write(f"#deadheadings: {deadheadings} | uncovered: {uncovered_tasks} \n")
                if line.startswith("The total"):
                    parts = line.split()
                    runtime = str(round(float(parts[4]),1))
                    runtime_matrix[(instance_name, window_size, runs_per_window, min_loco_knowledge, min_section_knowledge)] = runtime
                    #outfile.write(f"runtime: {runtime} \n")

            print(runtime_matrix)
            for window_size in window_size_list:
                outfile.write(f"window_size {window_size} \n")
                for runs_per_window in runs_per_window_list:
                    if runs_per_window == "1":
                        outfile.write(f"& {runs_per_window} ")
                    else:
                        outfile.write(f"& & {runs_per_window} ")
                    for min_loco_knowledge in min_loco_knowledge_list:
                        for min_section_knowledge in min_section_knowledge_list:
                            outfile.write(f"& {min_loco_knowledge}|{min_section_knowledge} ")
                            for instance_name in instance_list:
                                if (instance_name, window_size, runs_per_window, min_loco_knowledge, min_section_knowledge) in uncovered_task_matrix.keys() and (instance_name, window_size, runs_per_window, min_loco_knowledge, min_section_knowledge) in deadhead_matrix.keys() and (instance_name, window_size, runs_per_window, min_loco_knowledge, min_section_knowledge) in runtime_matrix.keys():
                                    outfile.write(f"& {uncovered_task_matrix[(instance_name, window_size, runs_per_window, min_loco_knowledge, min_section_knowledge)]} & {deadhead_matrix[(instance_name, window_size, runs_per_window, min_loco_knowledge, min_section_knowledge)]} & {runtime_matrix[(instance_name, window_size, runs_per_window, min_loco_knowledge, min_section_knowledge)]} ")
                                else:
                                    outfile.write(
                                        f"& - & - & - ")
                            outfile.write(" \\\ \n")

    infile.close()
    outfile.close()

def extract_duty_lengths(input_file, output_file):

    original_plan_read = False

    if input_file.endswith(".log"):
        with open(input_file, "r", encoding="utf-8") as infile:
             with open(output_file, "w", encoding="utf-8") as outfile:
                for line in infile:
                    if line.startswith("This are the duty categories of the original plan"):
                        if original_plan_read == False:
                            original_plan_duty_lengths  = [next(infile) for _ in range(4)]
                            original_plan_read = True
                        else:
                            rescheduled_plan_duty_lengths = [next(infile) for _ in range(4)]

                outfile.write(f"This are the duty categories of the original plan \n")
                for line in original_plan_duty_lengths:
                    outfile.write(line)
                outfile.write(f"This are the duty categories of the rescheduled plan \n")
                for line in rescheduled_plan_duty_lengths:
                    outfile.write(line)

def extract_solution_quality_sideways_table_excel_greedy(input_file, output_file):
    instance_list = []
    window_size_list = []
    runs_per_window_list = []
    method_list = []
    min_loco_knowledge_list = []
    min_section_knowledge_list = []
    max_deadhead_duration_list = []
    random_iteration_list = []

    nr_deadhead_matrix = {}
    deadhead_matrix = {}
    uncovered_task_matrix = {}
    runtime_matrix = {}
    spared_driver_matrix = {}

    # Parse input
    if input_file.endswith(".log"):
        with open(input_file, "r", encoding="utf-8") as infile:
            for line in infile:
                if line.startswith("Instance:"):
                    parts = line.split()
                    instance_name = parts[1].replace(".tsv", "")
                    if instance_name not in instance_list:
                        instance_list.append(instance_name)

                    window_size = parts[5]
                    if window_size not in window_size_list:
                        window_size_list.append(window_size)

                    runs_per_window = parts[10]
                    if runs_per_window not in runs_per_window_list:
                        runs_per_window_list.append(runs_per_window)

                    method = parts[13]
                    if method not in method_list:
                        method_list.append(method)

                    min_loco_knowledge = parts[18]
                    if min_loco_knowledge not in min_loco_knowledge_list:
                        min_loco_knowledge_list.append(min_loco_knowledge)

                    min_section_knowledge = parts[23]
                    if min_section_knowledge not in min_section_knowledge_list:
                        min_section_knowledge_list.append(min_section_knowledge)

                    max_deadhead_duration = parts[28]
                    if max_deadhead_duration not in max_deadhead_duration_list:
                        max_deadhead_duration_list.append(max_deadhead_duration)

                    random_iteration = parts[32]
                    if random_iteration not in random_iteration_list:
                        random_iteration_list.append(random_iteration)

                elif line.startswith("Initially the deadheading costs"):
                    parts = line.split()
                    deadheadings = str(int(int(parts[5].replace(",", ""))))
                    uncovered_tasks = parts[9]
                    nr_deadheads = parts[24]
                    spared_driver = parts[19]

                    nr_deadhead_matrix[(instance_name, window_size, runs_per_window, method,
                                     min_loco_knowledge, min_section_knowledge,
                                     max_deadhead_duration, random_iteration)] = nr_deadheads

                    deadhead_matrix[(instance_name, window_size, runs_per_window, method,
                                     min_loco_knowledge, min_section_knowledge,
                                     max_deadhead_duration, random_iteration)] = deadheadings

                    uncovered_task_matrix[(instance_name, window_size, runs_per_window, method,
                                           min_loco_knowledge, min_section_knowledge,
                                           max_deadhead_duration, random_iteration)] = uncovered_tasks

                    spared_driver_matrix[(instance_name, window_size, runs_per_window, method,
                                           min_loco_knowledge, min_section_knowledge,
                                           max_deadhead_duration, random_iteration)] = spared_driver

                elif line.startswith("In total there"):
                    parts = line.split()
                    runtime = str(round(float(parts[4]), 1))
                    runtime_matrix[(instance_name, window_size, runs_per_window, method,
                                    min_loco_knowledge, min_section_knowledge,
                                    max_deadhead_duration, random_iteration)] = runtime

    # Sort for consistent order
    instance_list.sort()
    window_size_list.sort()
    runs_per_window_list.sort(key=lambda x: int(x))

    avg_nr_deadhead_matrix = {}
    avg_deadhead_matrix = {}
    avg_uncovered_task_matrix = {}
    avg_runtime_matrix = {}
    avg_spared_driver_matrix = {}

    # Compute averages
    for instance_name in instance_list:
        for window_size in window_size_list:
            for runs_per_window in runs_per_window_list:
                for method in method_list:
                    for min_loco_knowledge in min_loco_knowledge_list:
                        for min_section_knowledge in min_section_knowledge_list:
                            for max_deadhead_duration in max_deadhead_duration_list:
                                nr_deadheads = 0
                                deadheads = 0
                                uncovered_tasks = 0
                                runtime = 0
                                spared_driver = 0
                                valid = True

                                for random_iteration in random_iteration_list:
                                    key = (instance_name, window_size, runs_per_window, method,
                                           min_loco_knowledge, min_section_knowledge,
                                           max_deadhead_duration, random_iteration)

                                    if key not in deadhead_matrix or \
                                       key not in uncovered_task_matrix or \
                                       key not in runtime_matrix:
                                        valid = False
                                        break

                                    nr_deadheads += float(nr_deadhead_matrix[key])
                                    deadheads += float(deadhead_matrix[key])
                                    uncovered_tasks += float(uncovered_task_matrix[key])
                                    runtime += float(runtime_matrix[key])
                                    spared_driver += float(spared_driver_matrix[key])

                                if valid:
                                    main_key = (instance_name, window_size, runs_per_window,
                                                method, min_loco_knowledge,
                                                min_section_knowledge, max_deadhead_duration)

                                    avg_nr_deadhead_matrix[main_key] = nr_deadheads / len(random_iteration_list)
                                    avg_deadhead_matrix[main_key] = deadheads / len(random_iteration_list)
                                    avg_uncovered_task_matrix[main_key] = uncovered_tasks / len(random_iteration_list)
                                    avg_runtime_matrix[main_key] = round(runtime / len(random_iteration_list), 2)
                                    avg_spared_driver_matrix[main_key] = spared_driver / len(random_iteration_list)

    # Create Excel files instead of logs
    for method in method_list:
        for min_loco_knowledge in min_loco_knowledge_list:
            for min_section_knowledge in min_section_knowledge_list:
                for max_deadhead_duration in max_deadhead_duration_list:

                    filename = output_file.replace(
                        ".log",
                        f"_{method}_{min_loco_knowledge}_{min_section_knowledge}_{max_deadhead_duration}.xlsx"
                    )

                    wb = Workbook()
                    ws = wb.active
                    ws.title = "Results"

                    # Header row
                    headers = [
                        "Instance",
                        "Characteristics",
                    ]

                    # DP windows & iterations
                    for window in ["150", "300", "450"]:
                        for it in ["1", "5", "10"]:
                            headers += [
                                f"DP {window}/{it} - Uncov",
                                f"DP {window}/{it} - Nr. Deadheads",
                                f"DP {window}/{it} - Deadhead Meters",
                                f"DP {window}/{it} - Runtime",
                                f"DP {window}/{it} - Nr. Spared Drivers"
                            ]

                    # DP full
                    for window in ["10800"]:
                        for it in ["1", "5", "10"]:
                            headers += [
                                f"DP FH/{it} - Uncov",
                                f"DP FH/{it} - Nr. Deadheads",
                                f"DP FH/{it} - Deadhead Meters",
                                f"DP FH/{it} - Runtime",
                                f"DP FH/{it} - Nr. Spared Drivers"
                            ]

                    # Model runs
                    for window in ["150", "300"]:
                        headers += [
                            f"Model {window} - Uncov",
                            f"Model {window} - Nr. Deadheads",
                            f"Model {window} - Deadhead Meters",
                            f"Model {window} - Runtime",
                            f"Model {window} - Nr. Spared Drivers"
                        ]

                    # model full
                    for window in ["10800"]:
                        headers += [
                            f"Model FH - Uncov",
                            f"Model FH - Nr. Deadheads",
                            f"Model FH - Deadhead Meters",
                            f"Model FH - Runtime",
                            f"Model FH - Nr. Spared Drivers"
                        ]

                    ws.append(headers)

                    # Fill table rows
                    for inst in instance_list:
                        row = [
                            inst,
                            "#tasks=...; #dh=...; dur=...; horizon=..."
                        ]

                        for window in ["150", "300", "450"]:
                            for it in ["1", "5", "10"]:
                                key = (inst, window, it, "DP",
                                       min_loco_knowledge, min_section_knowledge,
                                       max_deadhead_duration)

                                if key in avg_uncovered_task_matrix:
                                    row += [
                                        avg_uncovered_task_matrix[key],
                                        avg_nr_deadhead_matrix[key],
                                        avg_deadhead_matrix[key],
                                        avg_runtime_matrix[key],
                                        avg_spared_driver_matrix[key]
                                    ]
                                else:
                                    row += ["-", "-", "-", "-", "-"]

                        # Full horizon DP
                        for window in ["10800"]:
                            for it in ["1", "5", "10"]:
                                key = (inst, window, it, "DP",
                                       min_loco_knowledge, min_section_knowledge,
                                       max_deadhead_duration)

                                if key in avg_uncovered_task_matrix:
                                    row += [
                                        avg_uncovered_task_matrix[key],
                                        avg_nr_deadhead_matrix[key],
                                        avg_deadhead_matrix[key],
                                        avg_runtime_matrix[key],
                                        avg_spared_driver_matrix[key]
                                    ]
                                else:
                                    row += ["-", "-", "-", "-", "-"]

                        # Model results
                        for window in ["150", "300"]:
                            key = (inst, window, "1", "model",
                                   min_loco_knowledge, min_section_knowledge,
                                   max_deadhead_duration)

                            if key in avg_uncovered_task_matrix:
                                row += [
                                    avg_uncovered_task_matrix[key],
                                    avg_nr_deadhead_matrix[key],
                                    avg_deadhead_matrix[key],
                                    avg_runtime_matrix[key],
                                    avg_spared_driver_matrix[key]
                                ]
                            else:
                                row += ["-", "-", "-", "-", "-"]

                        # Model results
                        for window in ["10800"]:
                            key = (inst, window, "1", "model",
                                   min_loco_knowledge, min_section_knowledge,
                                   max_deadhead_duration)

                            if key in avg_uncovered_task_matrix:
                                row += [
                                    avg_uncovered_task_matrix[key],
                                    avg_nr_deadhead_matrix[key],
                                    avg_deadhead_matrix[key],
                                    avg_runtime_matrix[key],
                                    avg_spared_driver_matrix[key]
                                ]
                            else:
                                row += ["-", "-", "-", "-", "-"]

                        ws.append(row)

                    wb.save(filename)
